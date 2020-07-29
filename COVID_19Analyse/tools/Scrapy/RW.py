import abc
import os
import threading

import openpyxl


class RW(object):
    # 读
    @abc.abstractmethod
    def read(self, path, name):
        pass

    # 写
    @abc.abstractmethod
    def write(self, path, name, value):
        pass

    # 读取所有的sheet名字
    @abc.abstractmethod
    def readXLSX(self, path):
        pass


class ExcelData(RW):
    writeLock = threading.Lock()

    # {"英国"：[dict,dict,dict]}
    def writeDict_DictList(self, path, dict):

        workbook = openpyxl.Workbook()
        try:
            del workbook["Sheet"]
        except KeyError:
            pass
        try:
            for key in dict:
                sheet = workbook.create_sheet(key)
                value = self.__dictList2TwoDList(dict[key])
                for i in range(0, len(value)):
                    for j in range(0, len(value[i])):
                        sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            workbook.save(path)
        finally:
            workbook.close()

    def writeDict_2DList(self, path, dict):

        workbook = openpyxl.Workbook()
        try:
            del workbook["Sheet"]
        except KeyError:
            pass
        try:
            for key in dict:
                sheet = workbook.create_sheet(key)
                value = (dict[key])
                for i in range(0, len(value)):
                    for j in range(0, len(value[i])):
                        sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            workbook.save(path)
        finally:
            workbook.close()

    # 输入为一个二维list
    def write(self, path, sheet_name, value):
        with self.writeLock:
            index = len(value)
            if not os.path.exists(path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = sheet_name
            else:
                workbook = openpyxl.load_workbook(path)
                try:
                    del workbook[sheet_name]
                except KeyError:
                    pass
                sheet = workbook.create_sheet(sheet_name)

            try:
                for i in range(0, index):
                    for j in range(0, len(value[i])):
                        sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
                workbook.save(path)
                # print("xlsx格式表格写入数据成功！")
            finally:
                workbook.close()

    # 返回一个二维list
    def read(self, path, sheet_name):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        sheel = []
        for row in sheet.rows:
            rowtmp = []
            for cell in row:
                rowtmp.append(cell.value)
            sheel.append(rowtmp)
        # print(sheel)
        return sheel

    # 返回一堆workbook名字
    def readXLSX(self, path):
        wb = openpyxl.load_workbook(path)
        all_sheets = wb.get_sheet_names()
        return all_sheets

    # 功能：将一个dict list，转为二维列表,第一列为label
    # input：dict list
    # output：二维列表
    def __dictList2TwoDList(self, dictList):
        if len(dictList) < 1:
            return []
        title = []
        for key in dictList[0]:
            title.append(key)
        ans = [title]
        for dictTmp in dictList:
            Listtmp = []
            for key in dictTmp:
                Listtmp.append(dictTmp[key])
            ans.append(Listtmp)
        return ans
